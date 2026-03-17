import os
import sys
import smtplib
import schedule
import time
import datetime
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pymongo import MongoClient
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image

# Load environment variables
load_dotenv()

# --- Configuration ---
MONGO_URI = os.getenv("MONGO_URI", "mongodb://localhost:27017/steel_dms")
DB_NAME = "steel_dms" # Default from .env or override
REPORT_RECIPIENT = "hariithejj05@gmail.com"
EMAIL_SENDER = os.getenv("EMAIL_USER", "steel-dms-reports@example.com")
EMAIL_PASSWORD = os.getenv("EMAIL_PASS", "") # Needs to be set in .env
SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", 587))

# Path to logo (relative to backend)
LOGO_PATH = os.path.join(os.path.dirname(__file__), "../../../frontend/src/assets/excel_im/excel_img.png")

def get_project_stats():
    """Fetches project statistics from MongoDB."""
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    
    projects_coll = db["projects"]
    drawings_coll = db["drawings"]
    rfis_coll = db["rfis"]
    
    projects_data = []
    
    for proj in projects_coll.find():
        proj_id = proj["_id"]
        
        # Aggregating drawing stats
        drawing_count = drawings_coll.count_documents({"projectId": proj_id})
        
        # Approval vs Fabrication counts (approximation based on revision history)
        # Note: In a production system, these would be indexed counters
        approval_count = 0
        fabrication_count = 0
        
        drawings = drawings_coll.find({"projectId": proj_id})
        for dwg in drawings:
            fields = dwg.get("extractedFields", {})
            rev = str(fields.get("revision", "")).upper()
            if rev.isdigit():
                fabrication_count += 1
            elif rev and rev.isalpha():
                approval_count += 1
        
        # RFI Stats
        open_rfis = rfis_coll.count_documents({"projectId": proj_id, "status": "open"})
        closed_rfis = rfis_coll.count_documents({"projectId": proj_id, "status": "closed"})
        
        projects_data.append({
            "name": proj.get("name", "N/A"),
            "client": proj.get("clientName", "N/A"),
            "status": proj.get("status", "active"),
            "approximate": proj.get("approximateDrawingsCount", 0),
            "actual": drawing_count,
            "approval": approval_count,
            "fabrication": fabrication_count,
            "open_rfis": open_rfis,
            "closed_rfis": closed_rfis
        })
    
    client.close()
    return projects_data

def generate_report_excel(data):
    """Generates the Project Status Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Status"
    
    # --- Styling ---
    pink_fill = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    header_font = Font(bold=True, size=11)
    brand_font = Font(bold=True, size=14)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- Branding ---
    if os.path.exists(LOGO_PATH):
        try:
            img = Image(LOGO_PATH)
            # Scale if needed, but here we just place it
            ws.add_image(img, 'A1')
        except Exception as e:
            print(f"Error adding logo: {e}")

    # Spacer for logo (rows 1-5)
    for r in range(1, 7):
        ws.row_dimensions[r].height = 18
        
    # Brand Row
    ws.merge_cells('A7:L7')
    brand_cell = ws['A7']
    brand_cell.value = "CALDIM ENGINEERING PVT LTD - PROJECT STATUS REPORT"
    brand_cell.fill = pink_fill
    brand_cell.font = brand_font
    brand_cell.alignment = Alignment(horizontal="center", vertical="middle")
    ws.row_dimensions[7].height = 30
    
    # Date Row
    ws.merge_cells('I8:L8')
    date_cell = ws['I8']
    date_cell.value = f"Generated On: {datetime.date.today().strftime('%d/%m/%Y')}"
    date_cell.alignment = Alignment(horizontal="right")
    
    # Headers
    headers = [
        "Project Name", "Client/Fabricator", "Status", 
        "Approximate Drawings", "Extracted Drawings", 
        "Approved Count", "Fabrication Count",
        "Approval %", "Fabrication %",
        "Open RFIs", "Closed RFIs", "Overall Status"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num)
        cell.value = header
        cell.fill = yellow_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="middle", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[10].height = 30
    
    # Data
    for row_idx, proj in enumerate(data, 11):
        approx = proj["approximate"] or 1 # Avoid division by zero
        app_pct = round((proj["approval"] / approx) * 100)
        fab_pct = round((proj["fabrication"] / approx) * 100)
        
        row_data = [
            proj["name"], proj["client"], proj["status"].capitalize(),
            proj["approximate"], proj["actual"],
            proj["approval"], proj["fabrication"],
            f"{app_pct}%", f"{fab_pct}%",
            proj["open_rfis"], proj["closed_rfis"], "In Progress" if proj["status"] == "active" else proj["status"].capitalize()
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="middle")

    # Column Widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 20

    filename = f"Project_Status_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(os.path.dirname(__file__), "../../uploads/temp_reports", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def send_email(attachment_path):
    """Sends the report via email."""
    if not EMAIL_PASSWORD:
        print("Error: EMAIL_PASS not set in .env. Skipping email.")
        return False
        
    msg = MIMEMultipart()
    msg['From'] = EMAIL_SENDER
    msg['To'] = REPORT_RECIPIENT
    msg['Subject'] = "Weekly Project Status Report"
    
    body = f"""Hello,
    
Please find the attached Weekly Project Status Report for all currently active drawings.

Generation Date: {datetime.date.today().strftime('%B %d, %Y')}

Best Regards,
Steel DMS Automated System
"""
    msg.attach(MIMEText(body, 'plain'))
    
    filename = os.path.basename(attachment_path)
    try:
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f"attachment; filename= {filename}")
            msg.attach(part)
            
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            text = msg.as_string()
            server.sendmail(EMAIL_SENDER, REPORT_RECIPIENT, text)
            server.quit()
            print(f"Email sent successfully to {REPORT_RECIPIENT}")
            return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False

def job():
    """The main scheduled job."""
    print(f"[{datetime.datetime.now()}] Starting weekly report generation...")
    try:
        data = get_project_stats()
        report_path = generate_report_excel(data)
        send_email(report_path)
        print(f"[{datetime.datetime.now()}] Job completed.")
    except Exception as e:
        print(f"Error in scheduled job: {e}")

# --- Scheduler ---
def run_scheduler():
    # Schedule for Wednesday at 09:00 AM
    schedule.every().wednesday.at("09:00").do(job)
    
    print("Scheduler started. Waiting for next Wednesday...")
    while True:
        schedule.run_pending()
        time.sleep(60)

if __name__ == "__main__":
    # If run with --now flag, trigger immediately
    if "--now" in sys.argv:
        job()
    else:
        run_scheduler()
